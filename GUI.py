import cv2
import numpy as np
import tkinter as tk
import webbrowser

def preprocess(id):
    detector = cv2.CascadeClassifier('cascades/haarcascade_frontalface_default.xml')
    cam = cv2.VideoCapture(0)

    Id = id
    count = 0

    while (True):
        ret, img = cam.read()
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        faces = detector.detectMultiScale(gray, 1.3, 5)
        for (x, y, w, h) in faces:
            # incrementing sample number
            count = count + 1
            # saving the captured face in the dataset folder
            cv2.imwrite("dataset/User." + Id + '.' + str(count) + ".jpg", gray[y:y + h, x:x + w])  #
            cv2.rectangle(img, (x, y), (x + w, y + h), (255, 0, 0), 2)
            cv2.waitKey(100)
        cv2.imshow('face', img)
        cv2.waitKey(1)
        if (count > 40):
            break
    cam.release()
    cv2.destroyAllWindows()
    root = tk.Tk()
    canvas = tk.Canvas(root, height=100, width=300)
    canvas.pack()
    msg_lb = tk.Label(root, text='Successfully Registered Candidate Details')
    msg_lb.place(relx=0, rely=0.5, relheight=0.1, relwidth=1)

    root.mainloop()

def train():
    import cv2
    import os
    import numpy as np
    from PIL import Image

    recognizer = cv2.face.LBPHFaceRecognizer_create()
    path = 'dataset'

    def getImagesWithID(path):
        imagePaths = [os.path.join(path, f) for f in os.listdir(path)]
        faces = []
        IDs = []
        for imagePath in imagePaths:
            faceImg = Image.open(imagePath)
            faceNp = np.array(faceImg, 'uint8')
            ID = int(os.path.split(imagePath)[-1].split('.')[1])
            faces.append(faceNp)
            print(ID)
            IDs.append(ID)
            cv2.imshow("training", faceNp)
            cv2.waitKey(10)
        return IDs, faces

    Ids, faces = getImagesWithID(path)
    recognizer.train(faces, np.array(Ids))
    recognizer.save('trainningdata.yml')
    cv2.destroyAllWindows()
    root = tk.Tk()
    canvas = tk.Canvas(root, height=100, width=300)
    canvas.pack()
    msg_lb = tk.Label(root, text='Successfully Recognized')
    msg_lb.place(relx=0, rely=0.5, relheight=0.2, relwidth=1)

    root.mainloop()


def test():
    import cv2, os
    import numpy as np
    from PIL import Image

    recognizer = cv2.face.LBPHFaceRecognizer_create()
    recognizer.read("trainningdata.yml")
    cascadePath = "cascades/haarcascade_frontalface_default.xml"
    faceCascade = cv2.CascadeClassifier(cascadePath)

    cam = cv2.VideoCapture(0)
    font = cv2.FONT_HERSHEY_SIMPLEX
    while True:
        ret, im = cam.read()
        gray = cv2.cvtColor(im, cv2.COLOR_BGR2GRAY)
        faces = faceCascade.detectMultiScale(gray, 1.2)
        for (x, y, w, h) in faces:
            cv2.rectangle(im, (x, y), (x + w, y + h), (225, 0, 0), 2)
            Id, conf = recognizer.predict(gray[y:y + h, x:x + w])
            if (conf < 50):
                if (Id == 1):
                    var = 'sandeep'
                    Id = "sandeep {0:.2f}%".format(round(100 - conf, 2))
                elif (Id == 2):
                    var = 'rehan'
                    Id = "rehan {0:.2f}%".format(round(100 - conf, 2))
                elif (Id == 3):
                    var = 'pradeep'
                    Id = "pradeep {0:.2f}%".format(round(100 - conf, 2))
                elif (Id == 4):
                    var = 'rishab'
                    Id = "rishab {0:.2f}%".format(round(100 - conf, 2))
                elif (Id == 5):
                    var = 'meeta'
                    Id = "Meeta {0:.2f}%".format(round(100 - conf, 2))
                # elif (Id == new_id):
                #     var = 'candidate_name'
                #     Id = "candidate_name {0:.2f}%".format(round(100 - conf, 2))
            else:
                Id = "Unknown"
            cv2.putText(im, str(Id), (x, y - 10), font, 0.55, (0, 255, 0), 1)

        cv2.imshow('im', im)
        if cv2.waitKey(10) == ord('q'):
            break
    cam.release()
    cv2.destroyAllWindows()
    #create candidate row whenever you update new candidate
    import openpyxl
    import datetime

    curr = str(datetime.datetime.now())

    from openpyxl import Workbook
    from openpyxl.reader.excel import load_workbook

    book = Workbook()
    sheet = book.active

    #book.save("sample.xlsx")

    wb = load_workbook('sample.xlsx')
    ws = wb.get_sheet_by_name('Sheet')

    def write_excel(k):
        for i in range(3, ws.max_row + 1):
            if (ws.cell(row=i, column=3).value == var):
                ws.cell(row=i, column=k).value = 'P'
                break
        for i in range(3, ws.max_row + 1):
            if (ws.cell(row=i, column=k).value != 'P'):
                ws.cell(row=i, column=k).value = 'A'

    for i in range(3, ws.max_column + 2):
        if (ws.cell(row=2, column=i).value == curr[:11]):
            write_excel(i)
            break
        elif (ws.cell(row=2, column=i).value == None):
            ws.cell(row=2, column=i).value = curr[:11]
            write_excel(i)
            break

    wb.save('sample.xlsx')


def window2():
    rt = tk.Tk()
    canvas = tk.Canvas(rt, height=500, width=500, bg='#e6efef')
    canvas.pack()

    main_frame = tk.Frame(rt, bg='#e1f7f6',bd=5)
    main_frame.place(relx=0.1, rely=0.1, relheight=0.8, relwidth=0.8)

    id_lb = tk.Label(main_frame, text='CANDIDATE DETAILS')
    id_lb.place(relx=0.18, rely=0, relheight=0.1, relwidth=0.7)

    name_lb = tk.Label(main_frame, text='NAME : ')
    name_lb.place(relx=0, rely=0.1, relheight=0.1, relwidth=0.20)
    name_entry = tk.Entry(main_frame,bg='#ffffff')
    name_entry.place(relx=0.3,rely=0.1,relheight=0.1,relwidth=1)

    id_lb = tk.Label(main_frame, text='ID : ')
    id_lb.place(relx=0, rely=0.4, relheight=0.1, relwidth=0.20)
    id_entry = tk.Entry(main_frame, bg='#ffffff')
    id_entry.place(relx=0.3, rely=0.4, relheight=0.1, relwidth=1)

    sub_bt = tk.Button(main_frame,text='SUBMIT AND CAPTURE',command=lambda : preprocess(id_entry.get()))
    sub_bt.place(relx=0.4,rely=0.8,relheight=0.1,relwidth=0.50)
    rt.mainloop()

def upload():
    webbrowser.open('https://google.com')

#Main gui window and program starts from here
root = tk.Tk()

canvas = tk.Canvas(root,height = 500 , width=500,bg='#e6efef')
canvas.pack()

img = tk.PhotoImage(file="1.png")
bimg_lb = tk.Label(root,image = img)
bimg_lb.place(relheight=1,relwidth=1)

upl_bt = tk.Button(root, text = 'UPLOAD' , command =upload)
upl_bt.place(relx=0.7,rely=0.9,relheight=0.1,relwidth=0.25)

main_frame = tk.Frame(root,bg='#e1f7f6')
main_frame.place(relx=0.1,rely=0.1,relheight=0.8,relwidth=0.8)



new_bt = tk.Button(main_frame, text = 'NEW' , command =window2)
train_bt = tk.Button(main_frame, text = 'TRAIN',command=train)
test_bt = tk.Button(main_frame, text = 'TEST',command=test)
new_bt.place(relx=0,rely=0,relheight=0.3,relwidth=0.25)
train_bt.place(relx=0,rely=0.33,relheight=0.3,relwidth=0.25)
test_bt.place(relx=0,rely=0.66,relheight=0.3,relwidth=0.25)



ins1_lb = tk.Label(main_frame,text='Step 1 : capture candidate details')
ins1_lb.place(relx=0.25, rely=0, relheight=0.09, relwidth=0.6)

ins1_lb = tk.Label(main_frame,text='in proper lighting conditions')
ins1_lb.place(relx=0.25, rely=0.1, relheight=0.09, relwidth=0.6)

ins1_lb = tk.Label(main_frame,text='Step 2 : Train candidate images')
ins1_lb.place(relx=0.25, rely=0.35, relheight=0.09, relwidth=0.6)

ins1_lb = tk.Label(main_frame,text='Step 3 : Test candidate images')
ins1_lb.place(relx=0.25, rely=0.67, relheight=0.09, relwidth=0.6)

root.mainloop()


